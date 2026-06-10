"""Build a clean Excel workbook of all model results and a set of polished PNGs.

Produces under `lscm_milp/outputs/report/`:
    LSCM_Final_Results.xlsx        — data backbone (multiple sheets)
    pareto_frontier.png            — single chart, A4-portrait friendly
    carbon_price_emissions.png     — single chart
    carbon_price_modal_share.png   — single chart
    network_zone1.png              — vertical layout, hand-tuned labels
    network_zone2.png
    network_zone3.png
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import Alignment, Font, PatternFill

from . import data_loader, params as P
from .solver import (
    Config,
    carbon_price_sweep,
    deduplicate_pareto,
    detect_thresholds,
    epsilon_sweep,
    solve,
)

REPORT_DIR = P.OUTPUT_DIR / "report"
REPORT_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Hand-tuned label offsets so node labels do not collide on the small map
# ---------------------------------------------------------------------------

LABEL_OFFSETS = {
    # node_id : (dx, dy) in lon/lat units; (h_align, v_align) for placement.
    # Tuned so labels fan out from dense Yangtze / Sichuan / Pearl River clusters
    # without overlapping each other or the flow lines.
    #
    # Yangtze cluster (Suzhou supplier/prod, Shanghai DC, East+Central demand)
    "i2_Suzhou":   ( 4.0,  3.0, "left",   "bottom"),
    "j1_Suzhou":   ( 4.0,  1.5, "left",   "bottom"),
    "k1_Shanghai": ( 4.0,  0.0, "left",   "center"),
    "l1_East":     ( 4.0, -1.6, "left",   "top"),
    # Sichuan / West cluster (Chengdu DC, West demand, Chongqing prod)
    "k2_Chengdu":  (-3.5,  1.6, "right",  "bottom"),
    "l3_West":     (-3.5,  0.0, "right",  "center"),
    "j2_Chongqing":(-1.5, -2.0, "right",  "top"),
    # Pearl River cluster (Guangzhou supplier, Shenzhen DC, South demand)
    "i1_Guangzhou":(-3.5,  1.6, "right",  "bottom"),
    "l2_South":    (-3.5,  0.0, "right",  "center"),
    "k3_Shenzhen": ( 3.0, -1.6, "left",   "top"),
    # Bohai cluster (Tianjin supplier, North demand)
    "i3_Tianjin":  ( 3.5, -0.8, "left",   "top"),
    "l4_North":    ( 3.5,  1.4, "left",   "bottom"),
    # Central isolated
    "l5_Central":  ( 3.5,  0.0, "left",   "center"),
}

# Marker shape per node type — helps distinguish co-located supplier/production.
NODE_MARKER = {
    "supplier": "D",   # rhombus
    "prod":     "o",   # circle
    "dc":       "s",   # square
    "demand":   "^",   # triangle
}

NODE_DISPLAY = {
    "i1_Guangzhou": "Supplier · Guangzhou",
    "i2_Suzhou":   "Supplier · Yangtze",
    "i3_Tianjin":  "Supplier · Tianjin",
    "j1_Suzhou":   "Production · Suzhou",
    "j2_Chongqing":"Production · Chongqing",
    "k1_Shanghai": "DC · Shanghai",
    "k2_Chengdu":  "DC · Chengdu",
    "k3_Shenzhen": "DC · Shenzhen",
    "l1_East":     "Demand · East",
    "l2_South":    "Demand · South",
    "l3_West":     "Demand · West",
    "l4_North":    "Demand · North",
    "l5_Central":  "Demand · Central",
}

NODE_TYPE_COLOR = {
    "supplier": "#9aa0a6",
    "prod":     "#1f4e79",
    "dc":       "#c75450",
    "demand":   "#3e7a3a",
}

MODE_COLOR = {
    "road_diesel": "#c75450",
    "rail":        "#1f4e79",
    "nev_road":    "#3e7a3a",
}

MODE_LABEL = {
    "road_diesel": "Road diesel",
    "rail":        "Rail",
    "nev_road":    "NEV road",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def node_type(data, nid: str) -> str:
    if nid in data.suppliers: return "supplier"
    if nid in data.production: return "prod"
    if nid in data.dcs: return "dc"
    return "demand"


def is_active(data, nid: str, cfg: Config) -> bool:
    t = node_type(data, nid)
    if t == "supplier" or t == "demand":
        return True
    if t == "prod":
        return nid in cfg.open_sites
    return nid in cfg.open_dcs


# ---------------------------------------------------------------------------
# Excel workbook
# ---------------------------------------------------------------------------

HDR_FILL = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFFFF", size=11)
CENTER = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER


def _auto_width(ws, padding: int = 2):
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + padding, 40)


def build_workbook(data, pareto, sweep, jumps, sanity_results) -> Path:
    wb = Workbook()

    # ---- Cover sheet ----
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "From Compliance to Competitive Edge"
    cover["A1"].font = Font(bold=True, size=16, color="FF1F4E79")
    cover["A2"] = "Green Supply Chain Network Design Under China's Dual Carbon Goals"
    cover["A2"].font = Font(bold=True, size=12, color="FF1F4E79")
    cover["A4"] = "Companion data workbook"
    cover["A4"].font = Font(italic=True, size=11)
    cover["A6"] = "Author"
    cover["B6"] = "Alex Kuhne"
    cover["A7"] = "Course"
    cover["B7"] = "Logistics and Supply Chain Management (Prof. Zhang Jun, Tongji University)"
    cover["A8"] = "Generated"
    cover["B8"] = "2026-06-06"
    cover["A9"] = "Source code"
    cover["B9"] = "lscm_milp/ (Python package, PuLP MILP)"
    for r in range(6, 10):
        cover.cell(row=r, column=1).font = Font(bold=True, size=11)

    cover["A11"] = "Contents"
    cover["A11"].font = Font(bold=True, size=12, color="FF1F4E79")
    sheets_index = [
        ("Summary", "Headline scenario table with cost-minimum, first-threshold, "
                    "low-carbon and emissions-minimum configurations."),
        ("Pareto frontier", "Epsilon-constraint sweep (12 points) with cost, "
                            "emissions, modal shares, and an embedded chart."),
        ("Carbon price sweep", "Weighted-sum sweep across 41 carbon prices from "
                               "0 to 1000 RMB/tCO2, with two embedded charts."),
        ("Network nodes", "Coordinates of suppliers, production sites, DCs, and "
                          "demand regions, plus active status per zone."),
        ("Network flows", "Flow volumes and modes on every active arc in each zone."),
        ("Parameters", "Calibration block: capacities, emission factors, transport "
                       "costs, rail handling cost, distance thresholds."),
        ("Thresholds", "Structural-jump detection output from the carbon-price sweep."),
    ]
    for i, (sheet, desc) in enumerate(sheets_index):
        r = 12 + i
        cover.cell(row=r, column=1, value=sheet).font = Font(bold=True)
        cover.cell(row=r, column=2, value=desc)

    cover["A21"] = "Key result"
    cover["A21"].font = Font(bold=True, size=12, color="FF1F4E79")
    cover["A22"] = ("First structural threshold at p = 25 RMB/tCO2: modal shift "
                    "from 26% road / 74% rail to 11% / 89%, emissions -20%, cost "
                    "+0.02%. Current China ETS spot (Jan 2026) sits at 76 RMB/tCO2.")
    cover.merge_cells("A22:F22")
    cover["A22"].alignment = Alignment(wrap_text=True, vertical="top")
    cover.row_dimensions[22].height = 40

    cover.column_dimensions["A"].width = 20
    cover.column_dimensions["B"].width = 80

    # ---- Summary sheet ----
    ws = wb.create_sheet("Summary")
    ws["A1"] = "LSCM Final — MILP Results Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "First structural threshold"
    ws["B3"] = jumps[0]["carbon_price"] if jumps else None
    ws["C3"] = "RMB / tCO2"
    ws["A4"] = "Current China ETS spot (Jan 2026)"
    ws["B4"] = 76
    ws["C4"] = "RMB / tCO2"
    ws["A5"] = "Total demand modelled"
    ws["B5"] = sum(data.demand.values())
    ws["C5"] = "t / year"

    headers = [
        "Scenario", "Carbon price (RMB/tCO2)", "Cost (RMB/year)",
        "Emissions (tCO2/year)", "Δ cost vs base (%)",
        "Δ emissions vs base (%)", "Production sites", "DCs",
        "Road share", "Rail share", "NEV share",
    ]
    ws.append([])
    ws.append(["Representative configurations on the Pareto frontier"])
    ws["A8"].font = Font(bold=True, size=12)
    ws.append(headers)
    _style_header(ws, ws.max_row, len(headers))

    cm = pareto[-1]
    em = pareto[0]
    first_thr_cfg = None
    if jumps:
        for c in sweep:
            if c.carbon_price == jumps[0]["carbon_price"]:
                first_thr_cfg = c
                break
    low_carbon = pareto[1] if len(pareto) >= 2 else em

    rows = [
        ("Cost minimum",   cm),
        ("After first threshold (p=25)", first_thr_cfg),
        ("Low-carbon network (epsilon)", low_carbon),
        ("Emissions minimum (epsilon)",  em),
    ]
    for label, cfg in rows:
        if cfg is None:
            continue
        d_cost = (cfg.cost - cm.cost) / cm.cost * 100 if cm.cost else 0
        d_em   = (cm.emissions - cfg.emissions) / cm.emissions * 100 if cm.emissions else 0
        ws.append([
            label,
            cfg.carbon_price,
            round(cfg.cost, 2),
            round(cfg.emissions, 3),
            round(d_cost, 2),
            round(d_em, 2),
            " + ".join(cfg.open_sites),
            " + ".join(cfg.open_dcs),
            round(cfg.road_share, 4),
            round(cfg.rail_share, 4),
            round(cfg.nev_share, 4),
        ])

    _auto_width(ws)

    # ---- Pareto frontier sheet ----
    ws = wb.create_sheet("Pareto frontier")
    headers = ["Index", "Label", "Cost (M RMB/year)", "Emissions (tCO2/year)",
               "Open sites", "Open DCs", "Road share", "Rail share", "NEV share"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for i, cfg in enumerate(pareto):
        ws.append([
            i,
            cfg.objective_label,
            round(cfg.cost / 1e6, 3),
            round(cfg.emissions, 3),
            " + ".join(cfg.open_sites),
            " + ".join(cfg.open_dcs),
            round(cfg.road_share, 4),
            round(cfg.rail_share, 4),
            round(cfg.nev_share, 4),
        ])
    _auto_width(ws)

    # Embedded chart: Pareto frontier (cost vs emissions)
    chart = LineChart()
    chart.title = "Pareto frontier: cost vs emissions"
    chart.y_axis.title = "Emissions (tCO2 / year)"
    chart.x_axis.title = "Cost (M RMB / year)"
    chart.height = 12
    chart.width  = 18
    data_ref = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=len(pareto) + 1)
    cats     = Reference(ws, min_col=3, min_row=2, max_col=3, max_row=len(pareto) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "K1")

    # ---- Carbon price sweep sheet ----
    ws = wb.create_sheet("Carbon price sweep")
    headers = ["Carbon price (RMB/tCO2)", "Cost (M RMB/year)",
               "Emissions (tCO2/year)", "Road share", "Rail share", "NEV share",
               "Open sites", "Open DCs"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for cfg in sweep:
        ws.append([
            cfg.carbon_price,
            round(cfg.cost / 1e6, 3),
            round(cfg.emissions, 3),
            round(cfg.road_share, 4),
            round(cfg.rail_share, 4),
            round(cfg.nev_share, 4),
            " + ".join(cfg.open_sites),
            " + ".join(cfg.open_dcs),
        ])
    _auto_width(ws)

    # Embedded chart 1: emissions vs price
    n = len(sweep)
    chart = LineChart()
    chart.title = "Emissions vs carbon price"
    chart.y_axis.title = "Emissions (tCO2 / year)"
    chart.x_axis.title = "Carbon price (RMB / tCO2)"
    chart.height = 10; chart.width = 18
    em_ref = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=n + 1)
    cats   = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=n + 1)
    chart.add_data(em_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "J1")

    # Embedded chart 2: modal shares vs price (stacked)
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "percentStacked"
    chart2.overlap = 100
    chart2.title = "Modal share vs carbon price"
    chart2.y_axis.title = "Share of tonne-kilometres"
    chart2.x_axis.title = "Carbon price (RMB / tCO2)"
    chart2.height = 10; chart2.width = 18
    share_ref = Reference(ws, min_col=4, min_row=1, max_col=6, max_row=n + 1)
    chart2.add_data(share_ref, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, "J22")

    # ---- Nodes sheet ----
    ws = wb.create_sheet("Network nodes")
    headers = ["Node ID", "Type", "Name", "Longitude", "Latitude",
               "Active Zone 1", "Active Zone 2", "Active Zone 3"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    cfg_z1 = cm
    cfg_z2 = first_thr_cfg if first_thr_cfg else cm
    cfg_z3 = em
    for nid, (lon, lat) in data.coords.items():
        t = node_type(data, nid)
        ws.append([
            nid, t, NODE_DISPLAY.get(nid, nid), lon, lat,
            "yes" if is_active(data, nid, cfg_z1) else "no",
            "yes" if is_active(data, nid, cfg_z2) else "no",
            "yes" if is_active(data, nid, cfg_z3) else "no",
        ])
    _auto_width(ws)

    # ---- Flows sheet ----
    ws = wb.create_sheet("Network flows")
    headers = ["Zone", "Arc type", "Origin", "Destination", "Mode",
               "Flow (tonnes/year)", "Distance (km)"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for zone_label, cfg in (("Zone 1 cost min", cfg_z1),
                            ("Zone 2 post threshold", cfg_z2),
                            ("Zone 3 emissions min", cfg_z3)):
        for (atype, o, d, m), flow in cfg.flows.items():
            dist = (data.distance_rail if m == "rail" else data.distance_road)[(o, d)]
            ws.append([zone_label, atype, o, d, m, round(flow, 2), round(dist, 1)])
    _auto_width(ws)

    # ---- Parameters sheet ----
    ws = wb.create_sheet("Parameters")
    ws.append(["Parameter", "Value", "Unit", "Note"])
    _style_header(ws, 1, 4)
    rows = [
        ("Total demand",                P.TOTAL_DEMAND, "t/year", ""),
        ("Production capacity per site",P.CAPACITY_PRODUCTION, "t/year", ""),
        ("DC capacity per site",        P.CAPACITY_DC, "t/year", ""),
        ("Rail variable cost",          P.VARIABLE_COST["rail"], "RMB/tkm",
         "GLEC China range for containerised non-bulk freight"),
        ("Road diesel variable cost",   P.VARIABLE_COST["road_diesel"], "RMB/tkm", ""),
        ("NEV road variable cost",      P.VARIABLE_COST["nev_road"], "RMB/tkm", ""),
        ("Rail handling cost",          P.RAIL_HANDLING_COST_RMB_PER_T, "RMB/t",
         "Drayage + service-reliability premium"),
        ("Rail minimum distance",       P.RAIL_MIN_DISTANCE_KM, "km", ""),
        ("NEV maximum distance",        P.NEV_MAX_DISTANCE_KM, "km", ""),
        ("Emission factor road diesel", P.EMISSION_FACTORS["road_diesel"], "gCO2/tkm", "GLEC 2024"),
        ("Emission factor rail",        P.EMISSION_FACTORS["rail"], "gCO2/tkm", "GLEC 2024"),
        ("Emission factor NEV road",    P.EMISSION_FACTORS["nev_road"], "gCO2/tkm",
         "550 gCO2/kWh grid × 0.06 kWh/tkm"),
        ("Carbon price grid step",      25, "RMB/tCO2", ""),
        ("Carbon price grid max",       1000, "RMB/tCO2", ""),
    ]
    for r in rows:
        ws.append(list(r))
    _auto_width(ws)

    # ---- Threshold detection sheet ----
    ws = wb.create_sheet("Thresholds")
    ws.append(["Index", "Carbon price (RMB/tCO2)", "Reasons"])
    _style_header(ws, 1, 3)
    for i, j in enumerate(jumps):
        ws.append([i + 1, j["carbon_price"], "; ".join(j["reasons"])])
    _auto_width(ws)

    path = REPORT_DIR / "LSCM_Final_Results.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Matplotlib chart builders
# ---------------------------------------------------------------------------

# Common typography for A4-friendly figures
plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "axes.titlesize": 14,
    "axes.labelsize": 12,
    "xtick.labelsize": 10,
    "ytick.labelsize": 10,
    "legend.fontsize": 11,
    "figure.titlesize": 14,
})


def chart_pareto(pareto: list[Config], jumps: list[dict]) -> Path:
    path = REPORT_DIR / "pareto_frontier.png"
    fig, ax = plt.subplots(figsize=(6.5, 4.8))

    xs = [p.cost / 1e6 for p in pareto]
    ys = [p.emissions for p in pareto]
    ax.plot(xs, ys, "-", color="#1f4e79", linewidth=2.0, zorder=2)
    ax.scatter(xs, ys, s=36, color="#1f4e79", zorder=3)

    cm = pareto[-1]   # cost min — top of vertical branch
    em = pareto[0]    # em min — right end of horizontal step
    elbow = pareto[1] if len(pareto) >= 2 else cm

    x_min, x_max = min(xs), max(xs)
    y_min, y_max = min(ys), max(ys)
    pad_x = (x_max - x_min) * 0.18
    pad_y = (y_max - y_min) * 0.12
    ax.set_xlim(x_min - pad_x, x_max + pad_x)
    ax.set_ylim(y_min - pad_y, y_max + pad_y)

    # All four annotation boxes are stacked in the upper-right inside of the
    # L-shape — the natural whitespace area — with arrows pointing to the
    # corresponding data points on the left branch and the bottom-right tail.
    inner_x = x_min + (x_max - x_min) * 0.42
    top_y   = y_max - (y_max - y_min) * 0.05
    step    = (y_max - y_min) * 0.18

    box_blue = dict(boxstyle="round,pad=0.30", facecolor="white",
                    edgecolor="#1f4e79", alpha=0.95)
    box_red  = dict(boxstyle="round,pad=0.30", facecolor="white",
                    edgecolor="#c75450", alpha=0.95)
    box_grey = dict(boxstyle="round,pad=0.30", facecolor="white",
                    edgecolor="#666", alpha=0.95)

    ax.annotate(
        f"Cost minimum\n{cm.cost/1e6:.0f} M RMB · {cm.emissions:,.0f} tCO2",
        xy=(cm.cost/1e6, cm.emissions),
        xytext=(inner_x, top_y),
        fontsize=9, color="#1f4e79",
        ha="left", va="top",
        bbox=box_blue,
        arrowprops=dict(arrowstyle="->", color="#1f4e79", lw=0.9),
    )
    ax.annotate(
        "Zones 1–2\nmodal substitution\nnear-flat cost",
        xy=(elbow.cost/1e6, (cm.emissions + elbow.emissions) / 2),
        xytext=(inner_x, top_y - step),
        fontsize=9, color="#444",
        ha="left", va="top",
        bbox=box_grey,
        arrowprops=dict(arrowstyle="->", color="#666", lw=0.9),
    )
    z3_x = (elbow.cost/1e6 + em.cost/1e6) / 2
    z3_y = (elbow.emissions + em.emissions) / 2
    ax.annotate(
        "Zone 3\nfacility reconfiguration\nChongqing opens",
        xy=(z3_x, z3_y),
        xytext=(inner_x, top_y - 2 * step),
        fontsize=9, color="#444",
        ha="left", va="top",
        bbox=box_grey,
        arrowprops=dict(arrowstyle="->", color="#666", lw=0.9),
    )
    ax.annotate(
        f"Emissions minimum\n{em.cost/1e6:.0f} M RMB · {em.emissions:,.0f} tCO2",
        xy=(em.cost/1e6, em.emissions),
        xytext=(inner_x, top_y - 3 * step),
        fontsize=9, color="#c75450",
        ha="left", va="top",
        bbox=box_red,
        arrowprops=dict(arrowstyle="->", color="#c75450", lw=0.9),
    )

    ax.set_xlabel("Total cost (million RMB / year)")
    ax.set_ylabel("Total emissions (tCO2 / year)")
    ax.set_title("Pareto frontier — calibrated mini-case")
    ax.grid(True, alpha=0.3, linestyle="--")

    fig.tight_layout()
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)
    return path


def _add_broken_axis_marks(ax_left, ax_right, d: float = 0.012):
    """Draw the diagonal break marks on the inner spines of a broken x-axis."""
    kwargs = dict(transform=ax_left.transAxes, color="k", clip_on=False, lw=1.0)
    ax_left.plot((1 - d, 1 + d), (-d, +d), **kwargs)
    ax_left.plot((1 - d, 1 + d), (1 - d, 1 + d), **kwargs)
    kwargs.update(transform=ax_right.transAxes)
    ax_right.plot((-d, +d), (-d, +d), **kwargs)
    ax_right.plot((-d, +d), (1 - d, 1 + d), **kwargs)


def chart_carbon_price_emissions(sweep: list[Config], jumps: list[dict]) -> Path:
    path = REPORT_DIR / "carbon_price_emissions.png"
    # Broken x-axis: left zoomed on threshold transition, right shows plateau
    fig, (axL, axR) = plt.subplots(
        1, 2, sharey=True, figsize=(6.8, 4.0),
        gridspec_kw={"width_ratios": [3.2, 1], "wspace": 0.06},
    )
    xs = [c.carbon_price for c in sweep]
    ys = [c.emissions for c in sweep]

    for ax in (axL, axR):
        ax.plot(xs, ys, "-o", color="#1f4e79", linewidth=2.0, markersize=5)
        ax.grid(True, alpha=0.3, linestyle="--")

    axL.set_xlim(-5, 110)
    axR.set_xlim(890, 1015)

    # Hide the inner spines and the right ticks of the left panel
    axL.spines["right"].set_visible(False)
    axR.spines["left"].set_visible(False)
    axR.tick_params(left=False, labelleft=False)
    _add_broken_axis_marks(axL, axR)

    # Y range with a little headroom above the cost-min point
    y_top = max(ys) * 1.05
    y_bot = min(ys) * 0.97
    axL.set_ylim(y_bot, y_top)

    box_red = dict(boxstyle="round,pad=0.30", facecolor="white",
                   edgecolor="#c75450", alpha=0.95)
    box_grn = dict(boxstyle="round,pad=0.30", facecolor="white",
                   edgecolor="#3e7a3a", alpha=0.95)

    for j in jumps:
        p = j["carbon_price"]
        axL.axvline(p, color="#c75450", linestyle="--", linewidth=1.3, alpha=0.85)
        post = next((c for c in sweep if c.carbon_price == p), None)
        if post is not None:
            axL.annotate(
                f"Structural threshold\np = {p} RMB/tCO2\nrail 74% → 89%\nem 19,584 → 15,658",
                xy=(p, post.emissions),
                xytext=(45, 90), textcoords="offset points",
                fontsize=9, color="#c75450",
                arrowprops=dict(arrowstyle="->", color="#c75450", lw=0.9),
                ha="left", va="bottom",
                bbox=box_red,
            )

    # Current ETS spot — only on left panel (76 RMB sits inside left x-range)
    axL.axvline(76, color="#3e7a3a", linestyle=":", linewidth=1.4, alpha=0.9)
    axL.annotate(
        "China ETS spot\n(Jan 2026) ≈ 76",
        xy=(76, min(ys)),
        xytext=(8, 18), textcoords="offset points",
        fontsize=9, color="#3e7a3a",
        arrowprops=dict(arrowstyle="->", color="#3e7a3a", lw=0.9),
        ha="left", va="bottom",
        bbox=box_grn,
    )

    # Right panel label
    axR.text(0.5, 1.02, "plateau",
             transform=axR.transAxes,
             ha="center", va="bottom", fontsize=9, color="#555", style="italic")

    # Shared title and labels
    fig.suptitle("Carbon-price sweep: emissions trajectory",
                 fontsize=14, y=0.99)
    fig.supxlabel("Carbon price (RMB / tCO2)", fontsize=12, y=-0.02)
    axL.set_ylabel("Total emissions (tCO2 / year)")

    fig.tight_layout(rect=(0, 0, 1, 0.95))
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)
    return path


def chart_carbon_price_modal_share(sweep: list[Config], jumps: list[dict]) -> Path:
    path = REPORT_DIR / "carbon_price_modal_share.png"
    fig, (axL, axR) = plt.subplots(
        1, 2, sharey=True, figsize=(6.8, 4.0),
        gridspec_kw={"width_ratios": [3.2, 1], "wspace": 0.06},
    )
    xs = [c.carbon_price for c in sweep]
    road = [c.road_share for c in sweep]
    rail = [c.rail_share for c in sweep]
    nev  = [c.nev_share for c in sweep]

    for ax in (axL, axR):
        ax.stackplot(xs, road, rail, nev,
                     labels=[MODE_LABEL["road_diesel"], MODE_LABEL["rail"], MODE_LABEL["nev_road"]],
                     colors=[MODE_COLOR["road_diesel"], MODE_COLOR["rail"], MODE_COLOR["nev_road"]],
                     alpha=0.85)
        ax.grid(True, alpha=0.3, linestyle="--")

    axL.set_xlim(-5, 110)
    axR.set_xlim(890, 1015)
    axL.set_ylim(0, 1)

    axL.spines["right"].set_visible(False)
    axR.spines["left"].set_visible(False)
    axR.tick_params(left=False, labelleft=False)
    _add_broken_axis_marks(axL, axR)

    box_dark = dict(boxstyle="round,pad=0.30", facecolor="white",
                    edgecolor="#222", alpha=0.95)
    box_grn  = dict(boxstyle="round,pad=0.30", facecolor="white",
                    edgecolor="#3e7a3a", alpha=0.95)

    for j in jumps:
        p = j["carbon_price"]
        axL.axvline(p, color="#222", linestyle="--", linewidth=1.3, alpha=0.85)
        axL.annotate(
            f"Threshold\np = {p} RMB/tCO2\nroad 26% → 11%",
            xy=(p, 0.18),
            xytext=(35, 75), textcoords="offset points",
            fontsize=9, color="#222",
            arrowprops=dict(arrowstyle="->", color="#222", lw=0.9),
            ha="left", va="bottom",
            bbox=box_dark,
        )

    axL.axvline(76, color="#3e7a3a", linestyle=":", linewidth=1.4, alpha=0.95)
    axL.annotate(
        "ETS spot 76",
        xy=(76, 0.93),
        xytext=(8, -5), textcoords="offset points",
        fontsize=9, color="#3e7a3a",
        arrowprops=dict(arrowstyle="->", color="#3e7a3a", lw=0.9),
        ha="left", va="top",
        bbox=box_grn,
    )

    axR.text(0.5, 1.02, "plateau",
             transform=axR.transAxes,
             ha="center", va="bottom", fontsize=9, color="#555", style="italic")

    # Legend on right panel (less visual clutter on the zoomed left panel)
    axR.legend(loc="center left", bbox_to_anchor=(1.15, 0.5),
               framealpha=0.95, fontsize=9)

    fig.suptitle("Carbon-price sweep: modal share", fontsize=14, y=0.99)
    fig.supxlabel("Carbon price (RMB / tCO2)", fontsize=12, y=-0.02)
    axL.set_ylabel("Share of tonne-kilometres")

    fig.tight_layout(rect=(0, 0, 0.88, 0.95))
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)
    return path


def chart_network(data, cfg: Config, title: str, subtitle: str, filename: str) -> Path:
    path = REPORT_DIR / filename
    # Portrait body-image size for an A4 page with surrounding text — large
    # enough for readable labels, small enough not to dominate the page.
    fig, ax = plt.subplots(figsize=(5.2, 6.4))

    # Map extent — widen lon to make room for left/right-fanning labels.
    ax.set_xlim(92, 132)
    ax.set_ylim(17, 43)
    ax.set_aspect("auto")

    # Light gridlines, no axis box clutter
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)

    # Draw flows first (underneath)
    if cfg.flows:
        max_flow = max(cfg.flows.values())
        flow_order = sorted(
            cfg.flows.items(),
            key=lambda kv: {"road_diesel": 0, "rail": 1, "nev_road": 2}[kv[0][3]],
        )
        for (atype, o, d, m), flow in flow_order:
            x1, y1 = data.coords[o]
            x2, y2 = data.coords[d]
            lw = 1.0 + 3.5 * flow / max_flow
            ax.plot([x1, x2], [y1, y2],
                    color=MODE_COLOR[m], alpha=0.7,
                    linewidth=lw, zorder=2,
                    solid_capstyle="round")

    # Draw nodes (with type-specific markers)
    for nid, (lon, lat) in data.coords.items():
        t = node_type(data, nid)
        active = is_active(data, nid, cfg)
        marker_size = 230 if t in ("prod", "dc") else 180
        face = NODE_TYPE_COLOR[t]
        edge = "#222" if active else "#aaa"
        alpha = 1.0 if active else 0.30
        ax.scatter([lon], [lat], s=marker_size, c=face,
                   marker=NODE_MARKER[t],
                   edgecolor=edge, linewidth=1.3,
                   alpha=alpha, zorder=5)

    # Draw labels last with white background box for readability
    label_box = dict(boxstyle="round,pad=0.30", facecolor="white",
                     edgecolor="#bbb", alpha=0.92, linewidth=0.8)
    for nid, (lon, lat) in data.coords.items():
        active = is_active(data, nid, cfg)
        dx, dy, ha, va = LABEL_OFFSETS[nid]
        ax.annotate(
            NODE_DISPLAY[nid],
            xy=(lon, lat),
            xytext=(lon + dx, lat + dy),
            fontsize=9,
            ha=ha, va=va,
            alpha=1.0 if active else 0.55,
            color="#222" if active else "#666",
            arrowprops=dict(arrowstyle="-", color="#888", lw=0.7,
                            alpha=0.9 if active else 0.4),
            bbox=label_box,
            zorder=6,
        )

    # Headers
    fig.suptitle(title, fontsize=14, fontweight="bold", y=0.985)
    ax.set_title(subtitle, fontsize=10.5, color="#333", pad=10, loc="center")
    ax.set_xlabel("Longitude")
    ax.set_ylabel("Latitude")
    ax.grid(True, alpha=0.25, linestyle="--")

    # Legend: two rows. Node types on top, mode colours below.
    node_handles = [
        plt.Line2D([0], [0], marker=NODE_MARKER["supplier"], color="none",
                   markerfacecolor=NODE_TYPE_COLOR["supplier"], markeredgecolor="#222",
                   markersize=10, label="Supplier"),
        plt.Line2D([0], [0], marker=NODE_MARKER["prod"], color="none",
                   markerfacecolor=NODE_TYPE_COLOR["prod"], markeredgecolor="#222",
                   markersize=11, label="Production"),
        plt.Line2D([0], [0], marker=NODE_MARKER["dc"], color="none",
                   markerfacecolor=NODE_TYPE_COLOR["dc"], markeredgecolor="#222",
                   markersize=11, label="Distribution centre"),
        plt.Line2D([0], [0], marker=NODE_MARKER["demand"], color="none",
                   markerfacecolor=NODE_TYPE_COLOR["demand"], markeredgecolor="#222",
                   markersize=11, label="Demand region"),
    ]
    mode_handles = [
        mpatches.Patch(color=MODE_COLOR["road_diesel"], label=MODE_LABEL["road_diesel"]),
        mpatches.Patch(color=MODE_COLOR["rail"],        label=MODE_LABEL["rail"]),
        mpatches.Patch(color=MODE_COLOR["nev_road"],    label=MODE_LABEL["nev_road"]),
    ]
    leg1 = fig.legend(handles=node_handles, loc="lower center",
                      ncol=4, bbox_to_anchor=(0.5, 0.05),
                      frameon=False, fontsize=9,
                      handlelength=1.4, columnspacing=1.4)
    fig.legend(handles=mode_handles, loc="lower center",
               ncol=3, bbox_to_anchor=(0.5, 0.01),
               frameon=False, fontsize=9,
               handlelength=1.6, columnspacing=2.0)
    # add the first legend back
    fig.add_artist(leg1)

    fig.tight_layout(rect=(0, 0.08, 1, 0.97))
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading data ...")
    data = data_loader.load()

    print("Pareto sweep ...")
    pareto_raw = epsilon_sweep(data)
    pareto = deduplicate_pareto(pareto_raw)

    print("Carbon-price sweep ...")
    sweep = carbon_price_sweep(data)
    jumps = detect_thresholds(sweep)

    print("Sanity ...")
    sanity = []  # not needed for charts

    print("Writing Excel ...")
    xlsx = build_workbook(data, pareto, sweep, jumps, sanity)
    print(f"  {xlsx}")

    print("Building charts ...")
    print("  pareto", chart_pareto(pareto, jumps))
    print("  carbon emissions", chart_carbon_price_emissions(sweep, jumps))
    print("  carbon modal", chart_carbon_price_modal_share(sweep, jumps))

    # Identify Zone configurations
    cm = pareto[-1]
    em = pareto[0]
    post = None
    if jumps:
        for c in sweep:
            if c.carbon_price == jumps[0]["carbon_price"]:
                post = c
                break

    print("  network zone1",
          chart_network(
              data, cm,
              "Zone 1 — cost minimum",
              f"p = 0 RMB/tCO2 · cost 387 M RMB · 19,584 tCO2 · road 26% / rail 74%",
              "network_zone1.png",
          ))
    if post is not None:
        print("  network zone2",
              chart_network(
                  data, post,
                  "Zone 2 — after first structural threshold",
                  f"p = {post.carbon_price} RMB/tCO2 · cost 387 M RMB · 15,658 tCO2 · road 11% / rail 89%",
                  "network_zone2.png",
              ))
    print("  network zone3",
          chart_network(
              data, em,
              "Zone 3 — emissions minimum",
              "epsilon-constraint · cost 458 M RMB · 13,971 tCO2 · road 3% / rail 83% / NEV 15%",
              "network_zone3.png",
          ))

    print("Done. Outputs in", REPORT_DIR)


if __name__ == "__main__":
    main()
